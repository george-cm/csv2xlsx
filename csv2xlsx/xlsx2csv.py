import argparse
import csv
from pathlib import Path
from typing import List

from openpyxl import load_workbook


def parse_arguments():
    parser = argparse.ArgumentParser(
        prog="xlsx2csv", description="Convert a Excel xlsx file to csv."
    )
    parser.add_argument("input_xlsx_files", nargs="+")
    parser.add_argument("-sn", "--sheetnames", nargs="*")
    return parser.parse_args()


def fix_header_duplicate_fields(header):
    fields = list()
    new_header = list()
    for field in header:
        count = fields.count(field)
        if count == 0:
            new_header.append(field)
        else:
            new_header.append(f"{field}{count}")
        fields.append(field)
    return new_header


def xlsx2csv(
    xlsx_file: Path, sheet_names: List[str] | None = None
) -> List[Path | None]:
    written_csvs: List[Path] = []
    wb = load_workbook(xlsx_file.as_posix(), read_only=True, data_only=True)
    for sh_name in wb.sheetnames:
        sh = wb[sh_name]
        if (sheet_names is not None) and (sh.title not in sheet_names):
            continue
        csv_file = xlsx_file.parent / f"{xlsx_file.stem}_{sh.title}.csv"
        with csv_file.open("w", newline="", encoding="utf-8") as outf:
            writer = csv.writer(outf, dialect="excel")
            for i, row in enumerate(sh.rows):
                if i == 0:
                    header = [x.value for x in next(sh.rows)]
                    writer.writerow(fix_header_duplicate_fields(header))
                else:
                    line = [x.value for x in row]
                    writer.writerow(line)
        written_csvs.append(csv_file)
    return written_csvs


def main():
    args = parse_arguments()
    for xlsx in args.input_xlsx_files:
        xlsx_file = Path(xlsx)
        if not xlsx_file.exists():
            print(f"File does not exist: {xlsx_file.as_posix()}")
        if not xlsx_file.is_file():
            print(f"This is not a file: {xlsx_file.as_posix()}")
        xlsx2csv(xlsx_file, args.sheetnames)


if __name__ == "__main__":
    main()
