"""Tests for xlsx2csv module."""

import csv
import re
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from csv2xlsx.xlsx2csv import fix_header_duplicate_fields, xlsx2csv


@pytest.fixture
def tmp_xlsx(tmp_path: Path) -> Path:
    """Create a simple xlsx file with one sheet and known data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "London"])
    ws.append(["Bob", 25, "Paris"])
    ws.append(["Charlie", 35, "Berlin"])
    fpath = tmp_path / "test.xlsx"
    wb.save(fpath)
    wb.close()
    return fpath


@pytest.fixture
def tmp_xlsx_multi_sheet(tmp_path: Path) -> Path:
    """Create an xlsx file with two sheets."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "People"
    ws1.append(["Name", "Age"])
    ws1.append(["Alice", 30])

    ws2 = wb.create_sheet("Cities")
    ws2.append(["City", "Country"])
    ws2.append(["London", "UK"])
    ws2.append(["Paris", "France"])

    fpath = tmp_path / "multi.xlsx"
    wb.save(fpath)
    wb.close()
    return fpath


@pytest.fixture
def tmp_xlsx_duplicate_headers(tmp_path: Path) -> Path:
    """Create an xlsx file with duplicate column headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Name", "Name"])
    ws.append(["Alice", "Smith", "A."])
    fpath = tmp_path / "dupes.xlsx"
    wb.save(fpath)
    wb.close()
    return fpath


def _read_csv(path: Path) -> list[list[str]]:
    """Read a CSV file and return all rows as lists of strings."""
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.reader(f))


# --- fix_header_duplicate_fields ---


class TestFixHeaderDuplicateFields:
    def test_no_duplicates(self):
        assert fix_header_duplicate_fields(["A", "B", "C"]) == ["A", "B", "C"]

    def test_one_duplicate(self):
        assert fix_header_duplicate_fields(["A", "A"]) == ["A", "A1"]

    def test_multiple_duplicates(self):
        assert fix_header_duplicate_fields(["X", "X", "X"]) == ["X", "X1", "X2"]

    def test_mixed(self):
        assert fix_header_duplicate_fields(["A", "B", "A", "B", "A"]) == [
            "A", "B", "A1", "B1", "A2",
        ]

    def test_empty(self):
        assert fix_header_duplicate_fields([]) == []


# --- xlsx2csv basic conversion ---


class TestXlsx2CsvBasic:
    def test_returns_csv_paths(self, tmp_xlsx: Path):
        result = xlsx2csv(tmp_xlsx)
        assert len(result) == 1
        assert result[0].exists()
        assert result[0].suffix == ".csv"

    def test_csv_filename_includes_sheet_name(self, tmp_xlsx: Path):
        result = xlsx2csv(tmp_xlsx)
        assert result[0].name == "test_Sheet1.csv"

    def test_header_row_is_written(self, tmp_xlsx: Path):
        result = xlsx2csv(tmp_xlsx)
        rows = _read_csv(result[0])
        assert rows[0] == ["Name", "Age", "City"]

    def test_first_data_row_is_present(self, tmp_xlsx: Path):
        """The first data row must not be skipped."""
        result = xlsx2csv(tmp_xlsx)
        rows = _read_csv(result[0])
        assert len(rows) == 4  # 1 header + 3 data rows
        assert rows[1] == ["Alice", "30", "London"]

    def test_all_data_rows_are_present(self, tmp_xlsx: Path):
        result = xlsx2csv(tmp_xlsx)
        rows = _read_csv(result[0])
        assert rows[1] == ["Alice", "30", "London"]
        assert rows[2] == ["Bob", "25", "Paris"]
        assert rows[3] == ["Charlie", "35", "Berlin"]


# --- xlsx2csv with multiple sheets ---


class TestXlsx2CsvMultiSheet:
    def test_all_sheets_converted(self, tmp_xlsx_multi_sheet: Path):
        result = xlsx2csv(tmp_xlsx_multi_sheet)
        assert len(result) == 2
        names = {p.name for p in result}
        assert names == {"multi_People.csv", "multi_Cities.csv"}

    def test_filter_by_sheet_name(self, tmp_xlsx_multi_sheet: Path):
        result = xlsx2csv(tmp_xlsx_multi_sheet, sheet_names=["Cities"])
        assert len(result) == 1
        assert result[0].name == "multi_Cities.csv"

    def test_filter_nonexistent_sheet_returns_empty(self, tmp_xlsx_multi_sheet: Path):
        result = xlsx2csv(tmp_xlsx_multi_sheet, sheet_names=["NoSuchSheet"])
        assert result == []

    def test_filtered_sheet_data_is_correct(self, tmp_xlsx_multi_sheet: Path):
        result = xlsx2csv(tmp_xlsx_multi_sheet, sheet_names=["Cities"])
        rows = _read_csv(result[0])
        assert rows[0] == ["City", "Country"]
        assert rows[1] == ["London", "UK"]
        assert rows[2] == ["Paris", "France"]


# --- xlsx2csv with duplicate headers ---


class TestXlsx2CsvDuplicateHeaders:
    def test_duplicate_headers_are_disambiguated(
        self, tmp_xlsx_duplicate_headers: Path
    ):
        result = xlsx2csv(tmp_xlsx_duplicate_headers)
        rows = _read_csv(result[0])
        assert rows[0] == ["Name", "Name1", "Name2"]

    def test_data_preserved_with_duplicate_headers(
        self, tmp_xlsx_duplicate_headers: Path
    ):
        result = xlsx2csv(tmp_xlsx_duplicate_headers)
        rows = _read_csv(result[0])
        assert rows[1] == ["Alice", "Smith", "A."]


# --- xlsx2csv with bad dimension metadata ---


def _corrupt_dimension(xlsx_path: Path) -> None:
    """Rewrite the xlsx so the sheet dimension tag only declares A1,
    mimicking files exported by Microsoft Forms / SharePoint."""
    tmp = xlsx_path.with_suffix(".tmp")
    with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(tmp, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet"):
                text = data.decode("utf-8")
                text = re.sub(
                    r'<dimension ref="[^"]*"',
                    '<dimension ref="A1"',
                    text,
                )
                data = text.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(xlsx_path)


@pytest.fixture
def tmp_xlsx_bad_dimension(tmp_path: Path) -> Path:
    """Create an xlsx with multi-column data but a dimension tag claiming A1 only."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "Name", "Score"])
    ws.append([1, "Alice", 95])
    ws.append([2, "Bob", 87])
    fpath = tmp_path / "bad_dim.xlsx"
    wb.save(fpath)
    wb.close()
    _corrupt_dimension(fpath)
    return fpath


class TestXlsx2CsvBadDimension:
    def test_all_columns_are_present(self, tmp_xlsx_bad_dimension: Path):
        """Files with incorrect dimension metadata must still export all columns."""
        result = xlsx2csv(tmp_xlsx_bad_dimension)
        rows = _read_csv(result[0])
        assert rows[0] == ["ID", "Name", "Score"]

    def test_all_rows_are_present(self, tmp_xlsx_bad_dimension: Path):
        result = xlsx2csv(tmp_xlsx_bad_dimension)
        rows = _read_csv(result[0])
        assert len(rows) == 3  # 1 header + 2 data rows
        assert rows[1] == ["1", "Alice", "95"]
        assert rows[2] == ["2", "Bob", "87"]
